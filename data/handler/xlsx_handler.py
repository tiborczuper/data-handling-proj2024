from openpyxl import Workbook, load_workbook
from typing import List
from model_dataclasses import School, Passport, Person
from generator import generate_schools,generate_passport,generate_person


def write_schools_xlsx(filename: str, schools: List['School']):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

#header
    headers = ['SCHOOL_ID', 'SCHOOL_NAME', 'DISTRICT', 'SCHOOL_LEVEL', 'NUM_OF_STUDENTS']
    ws.append(headers)

    for school in schools:
        ws.append([school.school_id,school.name,school.district,school.school_level,school.students])
    wb.save(filename)

def write_passport_xlsx(filename: str, passports: List['Passport']):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

#header
    headers = ['PASSPORT_NUMBER','VALIDITY','VALID_FROM']
    ws.append(headers)

    for passport in passports:
        ws.append([passport.passport_number,passport.validity,passport.valid_from])
    wb.save(filename)

def write_person_xlsx(filename: str, persons: List['Person']):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

#header
    headers = ['ID_NUMBER', 'NAME', 'BIRTH_DATE', 'SCHOOL_NUMBER']
    ws.append(headers)

    for person in persons:
        ws.append([person.id_number,person.name,person.birth_date,person.school_number])
    wb.save(filename)

def read_schools_xlsx(filename: str):
    wb = load_workbook(filename)
    ws = wb.active
    ret = []

    for sor in ws.iter_rows(min_row=2, values_only=True):
        ret.append(School(sor[0],sor[1],sor[2],sor[3],sor[4]))

    return ret

def read_passports_xlsx(filename: str):
    wb = load_workbook(filename)
    ws = wb.active
    ret = []

    for sor in ws.iter_rows(min_row=2, values_only=True):
        ret.append(Passport(sor[0],sor[1],sor[2]))

    return ret

def read_persons_xlsx(filename: str):
    wb = load_workbook(filename)
    ws = wb.active
    ret = []

    for sor in ws.iter_rows(min_row=2, values_only=True):
        ret.append(Person(sor[0],sor[1],sor[2],sor[3]))

    return ret

if __name__ == '__main__':
    schools = generate_schools(3)
    passport = generate_passport(3)
    persons = generate_person(3,schools=schools,passports=passport)

    w_schools = write_schools_xlsx('schools.xlsx',schools)
    w_passport = write_passport_xlsx('passports.xlsx',passport)
    w_persons = write_person_xlsx('persons.xlsx',persons)

    r_schools = read_schools_xlsx('schools.xlsx')
    r_passport = read_passports_xlsx('passports.xlsx')
    r_persons = read_persons_xlsx('persons.xlsx')

    for v in r_schools:
        print(v)
    print()
    for v in r_passport:
        print(v)
    print()
    for v in r_persons:
        print(v)
